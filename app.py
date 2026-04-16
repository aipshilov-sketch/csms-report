"""
ЦСМС — Генератор отчёта
Автоматически формирует файл СОГЛАСОВАН из выгрузки биллинга.
"""

import json
import os
import re
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  Константы
# ─────────────────────────────────────────────
PRICE       = 70.10
VAT_FACTOR  = 22 / 122          # НДС включён в сумму
CLIENT_NAME = "Центр системы мониторинга рыболовства и связи"
REGISTRY_FILE = Path(sys.executable).parent / "registry.json" if getattr(sys, "frozen", False) \
                else Path(__file__).parent / "registry.json"

BLUE_HDR   = "1F3864"
LIGHT_BLUE = "BDD7EE"
GRAY_ROW   = "F2F2F2"
WHITE      = "FFFFFF"


# ─────────────────────────────────────────────
#  Справочник АТ → Филиал
# ─────────────────────────────────────────────
def load_registry() -> dict:
    """Загружает справочник {uid: филиал} из JSON."""
    if REGISTRY_FILE.exists():
        with open(REGISTRY_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_registry(registry: dict):
    """Сохраняет справочник на диск."""
    with open(REGISTRY_FILE, "w", encoding="utf-8") as f:
        json.dump(registry, f, ensure_ascii=False, indent=2)


# ─────────────────────────────────────────────
#  Парсинг примечаний
# ─────────────────────────────────────────────
def parse_gz_start(note: str, month: int, year: int) -> int | None:
    """
    Ищет 'ГЗ с ДД.ММ.ГГ' в примечании.
    Возвращает день начала если он совпадает с нужным месяцем/годом,
    иначе None (весь месяц).
    """
    if not note or str(note) == "nan":
        return None
    n = str(note)
    if "ВЫВОД" in n.upper():
        return None
    m = re.search(r"ГЗ\s+с\s+(\d{1,2})[.,/](\d{1,2})[.,/](\d{2,4})", n, re.IGNORECASE)
    if m:
        d, mo, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yr < 100:
            yr += 2000
        if mo == month and yr == year:
            return d
    return None


def parse_pdd_end(note: str, month: int, year: int) -> int | None:
    """
    Ищет 'ПДД с ДД.ММ.ГГ' в примечании.
    Возвращает день перехода (не включая) если совпадает с месяцем/годом.
    """
    if not note or str(note) == "nan":
        return None
    n = str(note)
    m = re.search(r"ПДД\s+с\s+(\d{1,2})[.,/](\d{1,2})[.,/](\d{2,4})", n, re.IGNORECASE)
    if m:
        d, mo, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yr < 100:
            yr += 2000
        if mo == month and yr == year:
            return d
    return None


# ─────────────────────────────────────────────
#  Обработка данных
# ─────────────────────────────────────────────
def process_source(path: str) -> tuple[pd.DataFrame, int, int]:
    """
    Читает исходный файл, фильтрует ЦСМС + только ГЗ-терминалы.
    Возвращает (df, month, year).
    """
    df = pd.read_excel(path)
    df["Примечание"] = df["Примечание"].astype(str)
    df["Описание"]   = df["Описание"].astype(str)

    # Только ЦСМС
    df = df[df["Клиент"] == CLIENT_NAME].copy()
    if df.empty:
        raise ValueError(f"Клиент «{CLIENT_NAME}» не найден в файле.")

    # ГЗ-терминалы: описание содержит ГЗ, либо терминал был ГЗ и перешёл на ПДД
    # (описание "ПДД В ATxxxx", но в примечании "ГЗ с ..." — учитываем дни до перехода)
    mask_gz = df["Описание"].str.contains("ГЗ", case=False, na=False)
    mask_pdd_was_gz = (
        df["Описание"].str.contains("ПДД", case=False, na=False) &
        df["Примечание"].str.contains(r"ГЗ\s+с", case=False, na=False)
    )
    df = df[mask_gz | mask_pdd_was_gz].copy()
    if df.empty:
        raise ValueError("Не найдено ни одной строки с ГЗ-терминалами.")

    df["Дата"] = pd.to_datetime(df["Дата"])
    df["day"]  = df["Дата"].dt.day

    # Определяем месяц/год из данных
    month = int(df["Дата"].dt.month.mode()[0])
    year  = int(df["Дата"].dt.year.mode()[0])

    return df, month, year


def build_uid_data(df: pd.DataFrame, month: int, year: int) -> dict:
    """
    Для каждого UID вычисляет:
      - активные дни с учётом ГЗ/ПДД ограничений
      - количество, сумму, НДС
    Возвращает {uid: {days, count, sum, vat, note}}
    """
    uid_notes = (
        df[df["Примечание"] != "nan"]
        [["UID", "Примечание"]]
        .drop_duplicates(subset="UID")
        .set_index("UID")["Примечание"]
        .to_dict()
    )

    result = {}
    for uid, group in df.groupby("UID"):
        note     = uid_notes.get(uid, "")
        gz_start = parse_gz_start(note, month, year) or 1
        pdd_end  = parse_pdd_end(note, month, year)  or 32

        src_days     = set(group["day"].unique())
        active_days  = {d for d in src_days if gz_start <= d < pdd_end}
        count        = len(active_days)
        total        = round(count * PRICE, 2)
        vat          = round(total * VAT_FACTOR, 2)

        result[uid] = {
            "days":  active_days,
            "count": count,
            "sum":   total,
            "vat":   vat,
            "note":  note,
        }

    return result


# ─────────────────────────────────────────────
#  Построение Excel
# ─────────────────────────────────────────────
def _thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_cell(cell, value, bg=BLUE_HDR, fg="FFFFFF", bold=True, size=9, wrap=True, align="center"):
    cell.value = value
    cell.font  = Font(bold=bold, color=fg, size=size, name="Arial")
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin_border()


def _data_cell(cell, value, bg=None, bold=False, num_fmt=None, align="center"):
    cell.value = value
    cell.font  = Font(bold=bold, size=9, name="Arial")
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _thin_border()
    if num_fmt:
        cell.number_format = num_fmt


def build_excel(uid_data: dict, registry: dict, month: int, year: int, out_path: str):
    """Строит итоговый Excel-файл в формате СОГЛАСОВАН."""

    # Имена месяцев
    MONTH_NAMES = {
        1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
        7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"
    }
    import calendar
    days_in_month = calendar.monthrange(year, month)[1]
    month_name    = MONTH_NAMES.get(month, str(month))

    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    # ── Строка 1: заголовок ──────────────────
    # Объединяем: A1:B1 пустые метки, C1 — период (охватывает дни), AH:AK — итоги
    ws.merge_cells(f"C1:{get_column_letter(2 + days_in_month)}1")
    _hdr_cell(ws["A1"], "№, п/п",  bg=BLUE_HDR)
    _hdr_cell(ws["B1"], "Сетевой № АТ", bg=BLUE_HDR)
    _hdr_cell(ws["C1"], f"{month_name}  {year}", bg=BLUE_HDR)

    col_count = 3 + days_in_month          # AH
    col_price = col_count + 1              # AI
    col_sum   = col_price + 1             # AJ
    col_vat   = col_sum + 1               # AK

    _hdr_cell(ws.cell(1, col_count), "Кол-во усл. ед за месяц",            bg=BLUE_HDR, wrap=True)
    _hdr_cell(ws.cell(1, col_price), "Стоимость одной условной единицы, руб", bg=BLUE_HDR, wrap=True)
    _hdr_cell(ws.cell(1, col_sum),   "Сумма всего, руб.",                   bg=BLUE_HDR, wrap=True)
    _hdr_cell(ws.cell(1, col_vat),   f"в т.ч. НДС (22%), руб",             bg=BLUE_HDR, wrap=True)
    ws.row_dimensions[1].height = 40

    # ── Строка 2: номера дней ────────────────
    ws["A2"].value = None
    ws["B2"].value = None
    for d in range(1, days_in_month + 1):
        cell = ws.cell(2, 2 + d)
        _hdr_cell(cell, d, bg=LIGHT_BLUE, fg="000000", bold=True, size=9)
    ws.row_dimensions[2].height = 18

    # ── Группируем АТ по филиалам ────────────
    # UIDs из данных, отсортированные; unknown = отдельная группа
    filials = {}
    unknown_uids = []
    for uid in sorted(uid_data.keys()):
        filial = registry.get(str(uid))
        if filial:
            filials.setdefault(filial, []).append(uid)
        else:
            unknown_uids.append(uid)

    if unknown_uids:
        filials["⚠ Филиал не указан"] = unknown_uids

    # ── Заполняем строки данных ──────────────
    current_row = 3
    seq_num     = 1
    data_rows   = []   # для формул итого

    for filial, uids in filials.items():
        # Строка-заголовок филиала
        ws.merge_cells(f"A{current_row}:{get_column_letter(col_vat)}{current_row}")
        cell = ws.cell(current_row, 1)
        cell.value = filial
        cell.font  = Font(bold=True, size=10, name="Arial", color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color="2E5D9E")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _thin_border()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for uid in uids:
            d = uid_data[uid]
            bg = WHITE if seq_num % 2 == 0 else GRAY_ROW

            _data_cell(ws.cell(current_row, 1), seq_num,  bg=bg, align="center")
            _data_cell(ws.cell(current_row, 2), uid,      bg=bg, bold=True, align="center")

            # Дни месяца
            for day in range(1, days_in_month + 1):
                cell = ws.cell(current_row, 2 + day)
                cell.value = "Р" if day in d["days"] else None
                cell.font  = Font(size=9, name="Arial",
                                  color="1F3864" if day in d["days"] else "AAAAAA",
                                  bold=(day in d["days"]))
                cell.fill  = PatternFill("solid", start_color=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = _thin_border()

            _data_cell(ws.cell(current_row, col_count), d["count"],  bg=bg, bold=True, align="center")
            _data_cell(ws.cell(current_row, col_price), PRICE,       bg=bg, num_fmt="#,##0.00", align="right")
            _data_cell(ws.cell(current_row, col_sum),   d["sum"],    bg=bg, num_fmt="#,##0.00", align="right")
            _data_cell(ws.cell(current_row, col_vat),   d["vat"],    bg=bg, num_fmt="#,##0.00", align="right")

            ws.row_dimensions[current_row].height = 15
            data_rows.append(current_row)
            seq_num     += 1
            current_row += 1

    # ── Строка ИТОГО ─────────────────────────
    itogo_row = current_row
    ws.merge_cells(f"A{itogo_row}:B{itogo_row}")
    cell = ws.cell(itogo_row, 1)
    cell.value = "Итого"
    cell.font  = Font(bold=True, size=10, name="Arial", color="FFFFFF")
    cell.fill  = PatternFill("solid", start_color=BLUE_HDR)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin_border()

    first_data = data_rows[0]
    last_data  = data_rows[-1]
    col_cnt_ltr   = get_column_letter(col_count)
    col_sum_ltr   = get_column_letter(col_sum)
    col_vat_ltr   = get_column_letter(col_vat)

    for col in [col_count, col_price, col_sum, col_vat]:
        ltr = get_column_letter(col)
        cell = ws.cell(itogo_row, col)
        if col == col_count:
            cell.value = f"=SUM({ltr}{first_data}:{ltr}{last_data})"
        elif col == col_price:
            cell.value = None
        elif col == col_sum:
            cell.value = f"=SUM({ltr}{first_data}:{ltr}{last_data})"
        elif col == col_vat:
            cell.value = f"=SUM({ltr}{first_data}:{ltr}{last_data})"
        cell.font  = Font(bold=True, size=10, name="Arial")
        cell.fill  = PatternFill("solid", start_color=LIGHT_BLUE)
        cell.alignment = Alignment(horizontal="center" if col == col_count else "right", vertical="center")
        cell.border = _thin_border()
        cell.number_format = "#,##0.00" if col in [col_sum, col_vat] else "0"

    ws.row_dimensions[itogo_row].height = 20

    # ── Ширина колонок ───────────────────────
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    for d in range(1, days_in_month + 1):
        ws.column_dimensions[get_column_letter(2 + d)].width = 3.2
    ws.column_dimensions[get_column_letter(col_count)].width = 10
    ws.column_dimensions[get_column_letter(col_price)].width = 14
    ws.column_dimensions[get_column_letter(col_sum)].width   = 14
    ws.column_dimensions[get_column_letter(col_vat)].width   = 14

    # Заморозка
    ws.freeze_panes = "C3"

    wb.save(out_path)
